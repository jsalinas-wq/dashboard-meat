#!/usr/bin/env python3
"""
Extracts raw rows from MASTER_SOCIAL_2026.xlsx (the MASTER_DATA and METAS 2026 sheets)
into the JSON shape netlify/functions/data.js serves and public/data-pipeline.js expects
— the same shape the Google Sheets API returns with UNFORMATTED_VALUE +
dateTimeRenderOption=SERIAL_NUMBER: a plain 2D array per sheet (header row + data rows),
with dates as day-serial numbers (days since 1899-12-30, matching Sheets/Excel).

This is intentionally "dumb": it does NOT aggregate or compute anything (no averages,
no goal matching). All of that logic lives in public/data-pipeline.js, runs in the
browser, and is already verified against this exact row shape — so this script only
needs to get the raw rows out of the spreadsheet correctly.

Usage:
    python3 rebuild_data.py <path-to-MASTER_SOCIAL_2026.xlsx> <output-dir>

Writes <output-dir>/master_data.json, <output-dir>/metas.json, <output-dir>/meta.json
(the last one just records the UTC rebuild timestamp, shown in the dashboard footer).
"""
import sys
import os
import json
import datetime

try:
    import openpyxl
except ImportError:
    print("openpyxl is required: pip install openpyxl --break-system-packages", file=sys.stderr)
    sys.exit(1)

SHEETS_EPOCH = datetime.datetime(1899, 12, 30)

# Bounded to the known schema widths — mirrors MASTER_RANGE/METAS_RANGE that used to be
# passed to the Google Sheets API in the live-fetch version of this project.
MASTER_MAX_COL = 14  # fecha_publicacion .. nuevos_seguidores
METAS_MAX_COL = 7    # Marca, Métrica, Facebook, Instagram, TikTok, LinkedIn, X


def cellval(v):
    if isinstance(v, datetime.datetime):
        delta = v - SHEETS_EPOCH
        return delta.days + delta.seconds / 86400
    return v


def extract_sheet(ws, max_col=None):
    kwargs = {"min_row": 1, "max_row": ws.max_row, "values_only": True}
    if max_col:
        kwargs["max_col"] = max_col
    return [[cellval(v) for v in row] for row in ws.iter_rows(**kwargs)]


def main():
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(1)
    xlsx_path, out_dir = sys.argv[1], sys.argv[2]
    os.makedirs(out_dir, exist_ok=True)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "MASTER_DATA" not in wb.sheetnames or "METAS 2026" not in wb.sheetnames:
        print(f"ERROR: esperaba las pestañas 'MASTER_DATA' y 'METAS 2026'; encontré: {wb.sheetnames}", file=sys.stderr)
        sys.exit(1)

    master_rows = extract_sheet(wb["MASTER_DATA"], MASTER_MAX_COL)
    metas_rows = extract_sheet(wb["METAS 2026"], METAS_MAX_COL)

    with open(os.path.join(out_dir, "master_data.json"), "w", encoding="utf-8") as f:
        json.dump(master_rows, f, ensure_ascii=False)
    with open(os.path.join(out_dir, "metas.json"), "w", encoding="utf-8") as f:
        json.dump(metas_rows, f, ensure_ascii=False)
    with open(os.path.join(out_dir, "meta.json"), "w", encoding="utf-8") as f:
        json.dump({"lastRebuiltAt": datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")}, f)

    print(f"OK: {len(master_rows)-1} filas de MASTER_DATA, {len(metas_rows)-1} filas de METAS 2026 -> {out_dir}")


if __name__ == "__main__":
    main()
